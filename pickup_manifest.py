# pickup_manifest.py

import os
import sys
import time
import traceback
import asyncio
import pytz
from datetime import datetime, timedelta
from pathlib import Path

from playwright.async_api import async_playwright
from xlrd import open_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

import sys
sys.stdout.reconfigure(encoding='utf-8')



START_HOUR = int(os.getenv("START_HOUR", "8"))
END_HOUR = int(os.getenv("END_HOUR", "23"))
FREQUENCY_MINUTES = int(os.getenv("FREQUENCY", "120"))
FACILITY_NAME = os.getenv("FACILITY_NAME", "PickUpManifest")
FOLDER_ID = os.getenv("FOLDER_ID", "FOLDER_ID")

SCRIPT_USERNAME = os.getenv("SCRIPT_USERNAME", "USERNAME")
SCRIPT_PASSWORD = os.getenv("SCRIPT_PASSWORD", "PASSWORD")


def authenticate_drive():
    creds = service_account.Credentials.from_service_account_file(
        "SERVICE_ACCOUNT_JSON.json",
        scopes=["https://www.googleapis.com/auth/drive", "https://www.googleapis.com/auth/spreadsheets"]
    )
    return build("drive", "v3", credentials=creds)

async def upload_to_drive(file_path: Path, folder_id: str):
    try:
        drive_service = authenticate_drive()
        file_metadata = {"name": file_path.name, "parents": [folder_id]}
        media = MediaFileUpload(str(file_path),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        uploaded_file = drive_service.files().create(
            body=file_metadata, media_body=media, fields="id"
        ).execute()
        print(f"[Pickup] ✅ Uploaded {file_path} (ID: {uploaded_file['id']})")
    except Exception as e:
        print(f"[Pickup] ❌ Error uploading {file_path}: {e}")
        traceback.print_exc()

def convert_xls_to_xlsx(xls_path: Path) -> Path:
    try:
        if not xls_path.exists():
            print(f"[Pickup] ❌ File not found: {xls_path}")
            return None

        xlsx_path = xls_path.with_suffix(".xlsx")
        print(f"[Pickup] Converting {xls_path} -> {xlsx_path}...")

        xls_book = open_workbook(str(xls_path), formatting_info=True)
        xlsx_book = Workbook()
        xlsx_book.remove(xlsx_book.active)

        for sheet_name in xls_book.sheet_names():
            xls_sheet = xls_book.sheet_by_name(sheet_name)
            new_sheet = xlsx_book.create_sheet(title=sheet_name)
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    new_sheet.cell(row=row_idx+1, column=col_idx+1,
                                   value=xls_sheet.cell_value(row_idx, col_idx))

            for (r1, r2, c1, c2) in xls_sheet.merged_cells:
                merge_range = f"{get_column_letter(c1+1)}{r1+1}:{get_column_letter(c2)}{r2}"
                new_sheet.merge_cells(merge_range)

        xlsx_book.save(str(xlsx_path))
        print(f"[Pickup] ✅ Converted: {xlsx_path}")
        return xlsx_path
    except Exception as e:
        print(f"[Pickup] ❌ XLS->XLSX conversion failed: {e}")
        traceback.print_exc()
        return None

async def run_pickup_iteration():
    """One iteration of the pickup-manifest logic."""
    est_tz = pytz.timezone("US/Eastern")
    new_date = datetime.now(est_tz).strftime("%m/%d/%Y")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()

        try:

            # Login sequence
            await page.goto("https://mybizaccount.fedex.com/")
            await page.click('input.credentials_input_submit')
            await page.wait_for_selector('#input28')
            await page.fill('#input28', SCRIPT_USERNAME)
            await page.fill('#input36', SCRIPT_PASSWORD)
            await page.click('input.button-primary')

            # Wait for 'domcontentloaded'
            await page.wait_for_load_state('domcontentloaded', timeout=60000)
            time.sleep(5)  # Consider replacing with an async sleep if possible
            # Search navigation
            await page.fill('input#PTSKEYWORD', 'FedEx Customer Connection')
            await page.click('#PTSSEARCHBTN')
            await page.wait_for_load_state('networkidle')
            await page.keyboard.press("Enter")
            await page.wait_for_load_state('networkidle')

            # P&D Manifest
            time.sleep(10)
            await page.wait_for_selector('#mainTabSettab_1', timeout=60000)
            await page.click('#mainTabSettab_1')
            time.sleep(15)
            await page.fill('#manifestForm\\:date_input', new_date)
            await page.click('#manifestForm\\:search')
            await asyncio.sleep(15)

            # Download
            async with page.expect_download() as dl_info:
                await page.click('#manifestForm\\:buttonGenerateExcel')
            dl = await dl_info.value

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            local_xls = Path.cwd() / f"{timestamp}_{FACILITY_NAME}.xls"
            await dl.save_as(local_xls)
            print(f"[Pickup] Downloaded: {local_xls}")

            # Convert
            xlsx_file = convert_xls_to_xlsx(local_xls)
            if xlsx_file:
                await upload_to_drive(xlsx_file, FOLDER_ID)
                if local_xls and local_xls.exists():
                    local_xls.unlink()
                if xlsx_file and xlsx_file.exists():
                    xlsx_file.unlink()

        finally:
            await browser.close()
            print("[Pickup] Browser closed")


def in_run_hours(now: datetime, start_hour: int, end_hour: int) -> bool:
    return (start_hour <= now.hour < end_hour)

def main():
    print(f"[Pickup] Starting with config: START_HOUR={START_HOUR}, END_HOUR={END_HOUR}, FREQ={FREQUENCY_MINUTES}min, FOLDER_ID={FOLDER_ID}")

    est_tz = pytz.timezone("US/Eastern")
    while True:
        try:
            now = datetime.now(est_tz)
            if in_run_hours(now, START_HOUR, END_HOUR):
                print(f"[Pickup] Within run hours ({START_HOUR}-{END_HOUR}), running iteration...")
                start_t = time.time()
                asyncio.run(run_pickup_iteration())
                elapsed = time.time() - start_t
                to_sleep = max(0, FREQUENCY_MINUTES*60 - elapsed)
                print(f"[Pickup] Sleeping {to_sleep:.0f}s until next iteration.")
                time.sleep(to_sleep)
            else:
                print(f"[Pickup] Outside run hours ({START_HOUR}-{END_HOUR}), waiting for next day...")
                next_run = now.replace(hour=START_HOUR, minute=0, second=0, microsecond=0)
                if next_run < now:
                    next_run += timedelta(days=1)
                sleep_secs = (next_run - now).total_seconds()
                print(f"[Pickup] Sleeping {sleep_secs/3600:.1f}h until {next_run}...")
                time.sleep(sleep_secs)

        except KeyboardInterrupt:
            print("[Pickup] KeyboardInterrupt => exiting.")
            sys.exit(0)
        except Exception as e:
            print(f"[Pickup] Unexpected main-loop error: {e}")
            traceback.print_exc()
            time.sleep(30)

if __name__ == "__main__":
    main()
