import asyncio
from playwright.async_api import async_playwright
from pathlib import Path
import time
from datetime import datetime, timedelta
import pytz
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import os

folder_id = 'FOLDER_ID'
import pandas as pd
import traceback
from pathlib import Path
import sys
# Existing utility functions
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from xlrd import open_workbook
import os


sys.stdout.reconfigure(encoding='utf-8')

def convert_xls_to_xlsx(xls_path) -> Path:
    """Convert an .xls file to .xlsx format while preserving merged formatting."""
    try:
        xls_path = Path(xls_path)  # Ensure xls_path is a Path object

        if not xls_path.exists():
            print(f"❌ File not found: {xls_path}")
            return None  # Avoid processing non-existent file

        xlsx_path = xls_path.with_suffix(".xlsx")

        print(f"🔄 Converting {xls_path} to {xlsx_path}...")

        # Read .xls file using xlrd
        xls_book = open_workbook(str(xls_path), formatting_info=True)
        xls_sheet = xls_book.sheet_by_index(0)

        # Create a new .xlsx workbook
        xlsx_book = Workbook()
        xlsx_sheet: Worksheet = xlsx_book.active

        # Copy data from .xls to .xlsx
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                xlsx_sheet.cell(row=row + 1, column=col + 1, value=xls_sheet.cell_value(row, col))

        # Preserve merged cells
        for merged_range in xls_sheet.merged_cells:
            r1, r2, c1, c2 = merged_range
            merge_range = f"{get_column_letter(c1 + 1)}{r1 + 1}:{get_column_letter(c2)}{r2}"
            xlsx_sheet.merge_cells(merge_range)

        # Save new .xlsx file
        xlsx_book.save(str(xlsx_path))
        print(f"✅ Successfully converted {xls_path} to {xlsx_path}")
        return xlsx_path

    except Exception as e:
        print(f"❌ Error converting {xls_path} to XLSX: {e}")
        traceback.print_exc()
        return None
def authenticate_drive():
    creds = service_account.Credentials.from_service_account_file(
        "Service Account JSON.json", 
        scopes=["https://www.googleapis.com/auth/drive", "https://www.googleapis.com/auth/spreadsheets"]
    )
    return build("drive", "v3", credentials=creds)

# Upload file to Google Drive
import traceback

async def upload_to_drive(file_path: Path, folder_id: str):
    """Upload an XLSX file to Google Drive."""
    try:
        drive_service = authenticate_drive()

        file_metadata = {
            "name": file_path.name,  # File name in Google Drive
            "parents": [folder_id],  # Folder ID in Google Drive
        }
        media = MediaFileUpload(
            str(file_path), 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"  # Correct MIME type for .xlsx
        )

        uploaded_file = drive_service.files().create(
            body=file_metadata, media_body=media, fields="id"
        ).execute()

        print(f"✅ Uploaded {file_path} to Google Drive with ID: {uploaded_file['id']}")

    except Exception as e:
        print(f"❌ Error uploading {file_path}: {e}")
        traceback.print_exc()



async def rename_csv(file_path: Path, facility_name: str):
    """Renames the downloaded CSV file to include timestamp and facility name."""
    # Get the current time in Eastern Standard Time (EST)
    est_tz = pytz.timezone('US/Eastern')
    current_time_est = datetime.now(est_tz)

    # Format the timestamp as YYYYMMDD_HHMMSS
    timestamp = current_time_est.strftime("%Y%m%d_%H%M%S")
    
    # Create the new file name
    new_name = f"{timestamp}_{facility_name}.xls"
    new_path = file_path.parent / new_name

    try:
        file_path.rename(new_path)
        print(f"Renamed file to: {new_path}")
        return new_path  # Return the new path for further use
    except Exception as e:
        print(f"Error renaming file: {e}")
        return None

import asyncio
import time

async def run_script_if_in_time_range():
    """ Runs the script every hour between 9 AM and 10 PM EST, printing time left every minute. """
    est_tz = pytz.timezone("US/Eastern")

    while True:
        now = datetime.now(est_tz)
        current_hour = now.hour
        if 9 <= current_hour < 22:  # Run only between 9 AM - 10 PM
            print(f"🟢 Running script at {now.strftime('%Y-%m-%d %I:%M %p EST')}")
            await automation_function()

            # Count down in minutes until the next hour
            for remaining in range(60, 0, -1):
                now = datetime.now(est_tz)
                print(f"⏳ Next run in {remaining} minutes... ({now.strftime('%I:%M %p EST')})")
                time.sleep(60)  # Wait for 1 minute before updating

        else:
            # Calculate sleep time until next 9 AM
            next_run = now.replace(hour=9, minute=0, second=0, microsecond=0) + timedelta(days=1)
            sleep_time = (next_run - now).total_seconds()
            print(f"⏸️ Off-hours! Sleeping until 9 AM EST ({next_run.strftime('%I:%M %p EST')})")
            time.sleep(sleep_time)  # Sleep until 9 AM

async def automation_function():
    # Get the current time in Eastern Standard Time (EST)
    est_tz = pytz.timezone('US/Eastern')
    current_time_est = datetime.now(est_tz)
    script_password = os.getenv("SCRIPT_PASSWORD")
    if script_password:
        print(f"✅ Successfully fetched password: {script_password}")
    else:
        print("❌ SCRIPT_PASSWORD environment variable is not set!")
    # Format the date as MM/DD/YYYY
    new_date = current_time_est.strftime("%m/%d/%Y")
    print(new_date)
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()

        # Login sequence
        await page.goto("https://mybizaccount.fedex.com/")
        await page.click('input.credentials_input_submit')
        await page.wait_for_selector('#input28')
        await page.fill('#input28', 'USERNAME')
        await page.fill('#input36', f'{script_password}')
        await page.click('input.button-primary')
        await page.wait_for_load_state('networkidle')

        # Search navigation
        await page.fill('input#PTSKEYWORD', 'Daily Service Wk & Vision IBPR')
        await page.click('#PTSSEARCHBTN')
        await page.wait_for_load_state('networkidle')
        await page.keyboard.press("Enter")
        await page.wait_for_load_state('networkidle')

        # Date input
        await page.fill('div#dateTimePicker1 input', new_date)
        
        while True:
            # Process current facility
            current_facility = await page.locator('select#facilitySelect option:checked').text_content()
            current_facility = current_facility.strip()
            print(f"Processing: {current_facility}")

            # Handle search
            try:
                await page.click('button.selectionButton')
                await page.wait_for_load_state('networkidle', timeout=60000)

                # Download handling
                excel_locator = page.locator("img[alt='Excel']")
                if await excel_locator.count() > 0:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    async with page.expect_download() as download_info:
                        await excel_locator.click()
                    download = await download_info.value

                    # Save the file with .xls extension
                    safe_name = current_facility.replace('/', '_').replace(' ', '_')
                    original_path = Path.cwd() / f"{safe_name}.xls"
                    await download.save_as(original_path)
                    print(f"Downloaded {original_path}")

                    # Optional: You can add logic to rename or convert the file if needed
                    renamed_path = await rename_csv(original_path, safe_name)
                    xlsx_path = await convert_xls_to_xlsx(renamed_path)
                    if xlsx_path:
                        # Upload the .xlsx file to Google Drive
                        await upload_to_drive(xlsx_path, folder_id)
                    else:
                        print(f"❌ Failed to rename file: {original_path}")
                else:
                    print(f"No Excel icon found for {current_facility}")
            except Exception as e:
                print(e)



            current_facility = 'ZECA-278'
            # Refill the date field
            print(f"Processing: {current_facility}")

            await page.fill('div#dateTimePicker1 input', new_date)

            # Simulate tab, enter, up arrow, and enter
            await page.keyboard.press("Tab")
            await page.keyboard.press("Enter")
            await page.keyboard.press("ArrowUp")
            await page.keyboard.press("Enter")
            time.sleep(5)
            # Handle search
            try:
                await page.click('button.selectionButton')
                await page.wait_for_load_state('networkidle', timeout=60000)

                # Download handling
                excel_locator = page.locator("img[alt='Excel']")
                if await excel_locator.count() > 0:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    async with page.expect_download() as download_info:
                        await excel_locator.click()
                    download = await download_info.value

                    # Save the file with .xls extension
                    safe_name = current_facility.replace('/', '_').replace(' ', '_')
                    original_path = Path.cwd() / f"{safe_name}.xls"
                    await download.save_as(original_path)
                    print(f"Downloaded {original_path}")

                    # Optional: You can add logic to rename or convert the file if needed
                    renamed_path = await rename_csv(original_path, safe_name)
                    xlsx_path = await convert_xls_to_xlsx(renamed_path)
                    if xlsx_path:
                        # Upload the .xlsx file to Google Drive
                        await upload_to_drive(xlsx_path, folder_id)
                    else:
                        print(f"❌ Failed to rename file: {original_path}")
                else:
                    print(f"No Excel icon found for {current_facility}")
            except Exception as e:
                print(e)




            current_facility = 'ZWLN-256/3256'
             # Refill the date field
            await page.fill('div#dateTimePicker1 input', new_date)
            print(f"Processing: {current_facility}")

            # Simulate tab, enter, up arrow, and enter
            await page.keyboard.press("Tab")
            await page.keyboard.press("Enter")
            await page.keyboard.press("ArrowDown")
            await page.keyboard.press("ArrowDown")
            await page.keyboard.press("Enter")
            time.sleep(5)
            # Handle search
            try:
                await page.click('button.selectionButton')
                await page.wait_for_load_state('networkidle', timeout=60000)

                # Download handling
                excel_locator = page.locator("img[alt='Excel']")
                if await excel_locator.count() > 0: 
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    async with page.expect_download() as download_info:
                        await excel_locator.click()
                    download = await download_info.value

                    # Save the file with .xls extension
                    safe_name = current_facility.replace('/', '_').replace(' ', '_')
                    original_path = Path.cwd() / f"{safe_name}.xls"
                    await download.save_as(original_path)
                    print(f"Downloaded {original_path}")

                    # Optional: You can add logic to rename or convert the file if needed
                    renamed_path = await rename_csv(original_path, safe_name)
                    xlsx_path = await convert_xls_to_xlsx(renamed_path)
                    if xlsx_path:
                        # Upload the .xlsx file to Google Drive
                        await upload_to_drive(xlsx_path, folder_id)
                    else:
                        print(f"❌ Failed to rename file: {original_path}")
                else:
                    print(f"No Excel icon found for {current_facility}")
            except Exception as e:
                print(e)
            break

    await browser.close()

asyncio.run(run_script_if_in_time_range())
