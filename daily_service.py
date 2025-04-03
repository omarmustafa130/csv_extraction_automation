# daily_service.py

import os
import sys
import time
import traceback
import asyncio
import pytz
from datetime import datetime, timedelta
from pathlib import Path

from playwright.async_api import async_playwright
from xlrd import open_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import sys
sys.stdout.reconfigure(encoding='utf-8')


# Read config from environment (passed by control_panel.py)
START_HOUR = int(os.getenv("START_HOUR", "9"))
END_HOUR = int(os.getenv("END_HOUR", "22"))
FREQUENCY_MINUTES = int(os.getenv("FREQUENCY", "60"))
FACILITY_NAME = os.getenv("FACILITY_NAME", "DailyService")
FOLDER_ID = os.getenv("FOLDER_ID", "NONE")

# Hard-coded credentials (or read from env if you prefer)
SCRIPT_USERNAME = os.getenv("SCRIPT_USERNAME", "DEFAULT-USERNAME")
SCRIPT_PASSWORD = os.getenv("SCRIPT_PASSWORD", "DEFAULT-PASSWORD")


FACILITIES = [
    {"name": "ZECA-278", "steps": []},
    {"name": "ZNHI-250/3250",      "steps": ["Tab", "Enter", "ArrowDown", "Enter"]},
    {"name": "ZWLN-256/3256", "steps": ["Tab", "Enter", "ArrowDown", "ArrowDown", "Enter"]},
]

def authenticate_drive():
    creds = service_account.Credentials.from_service_account_file(
        "SERVICE-JSON.json",
        scopes=["https://www.googleapis.com/auth/drive", "https://www.googleapis.com/auth/spreadsheets"]
    )
    return build("drive", "v3", credentials=creds)

async def upload_to_drive(file_path: Path, folder_id: str):
    try:
        drive_service = authenticate_drive()
        file_metadata = {"name": file_path.name, "parents": [folder_id]}
        media = MediaFileUpload(str(file_path),
                                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        uploaded_file = drive_service.files().create(
            body=file_metadata, media_body=media, fields="id"
        ).execute()
        print(f"[Daily] ✅ Uploaded {file_path} (ID: {uploaded_file['id']})")
    except Exception as e:
        print(f"[Daily] ❌ Error uploading {file_path}: {e}")
        traceback.print_exc()

def convert_xls_to_xlsx(xls_path: Path) -> Path:
    """Convert .xls to .xlsx while preserving merges."""
    try:
        if not xls_path.exists():
            print(f"[Daily] ❌ File not found: {xls_path}")
            return None

        xlsx_path = xls_path.with_suffix(".xlsx")
        print(f"[Daily] Converting {xls_path} -> {xlsx_path}...")

        xls_book = open_workbook(str(xls_path), formatting_info=True)
        xlsx_book = Workbook()
        xlsx_book.remove(xlsx_book.active)

        for sheet_name in xls_book.sheet_names():
            xls_sheet = xls_book.sheet_by_name(sheet_name)
            new_sheet = xlsx_book.create_sheet(title=sheet_name)
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    new_sheet.cell(row=row_idx+1, column=col_idx+1,
                                   value=xls_sheet.cell_value(row_idx, col_idx))

            for (r1, r2, c1, c2) in xls_sheet.merged_cells:
                merge_range = f"{get_column_letter(c1+1)}{r1+1}:{get_column_letter(c2)}{r2}"
                new_sheet.merge_cells(merge_range)

        xlsx_book.save(str(xlsx_path))
        print(f"[Daily] ✅ Converted to: {xlsx_path}")
        return xlsx_path
    except Exception as e:
        print(f"[Daily] ❌ XLS->XLSX conversion failed: {e}")
        traceback.print_exc()
        return None

async def run_daily_iteration():
    """One iteration of the daily logic (login, for each facility download...)."""
    est_tz = pytz.timezone("US/Eastern")
    new_date = datetime.now(est_tz).strftime("%m/%d/%Y")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()

        try:
            # Login sequence
            await page.goto("https://mybizaccount.fedex.com/")
            time.sleep(5)
            await page.click('input.credentials_input_submit')
            time.sleep(5)
            await page.wait_for_selector('#input28')
            await page.fill('#input28', SCRIPT_USERNAME)
            await page.fill('#input36', SCRIPT_PASSWORD)
            await page.click('input.button-primary')
            # Wait for 'domcontentloaded'
            await page.wait_for_load_state('domcontentloaded', timeout=60000)
            await page.wait_for_load_state('networkidle', timeout=60000)
            time.sleep(20)  # Consider replacing with an async sleep if possible
            # Search navigation
            page.wait_for_function("document.readyState === 'complete'")

            await page.fill('input#PTSKEYWORD', 'Daily Service Wk & Vision IBPR', timeout=60000)
            await page.click('#PTSSEARCHBTN')
            await page.wait_for_load_state('networkidle')
            time.sleep(5)
            await page.keyboard.press("Enter")
            await page.wait_for_load_state('networkidle')
            time.sleep(5)
            # Set the date
            await page.fill('div#dateTimePicker1 input', new_date)
            time.sleep(15)
            # For each facility
            for fac in FACILITIES:
                fac_name = fac["name"]
                print(f"[Daily] Processing facility: {fac_name}")

                # Re-fill date (sometimes it resets)
                await page.fill('div#dateTimePicker1 input', new_date)
                for key_step in fac["steps"]:
                    await page.keyboard.press(key_step)
                    await asyncio.sleep(0.5)

                try:
                    await page.click('button.selectionButton')
                    await page.wait_for_load_state('networkidle', timeout=60000)

                    excel_locator = page.locator("img[alt='Excel']")
                    if (await excel_locator.count()) > 0:
                        async with page.expect_download() as dl_info:
                            await excel_locator.click()
                        dl = await dl_info.value

                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        safe_name = fac_name.replace("/", "_").replace(" ", "_")
                        local_xls = Path.cwd() / f"{timestamp}_{safe_name}.xls"
                        await dl.save_as(local_xls)
                        print(f"[Daily] Downloaded: {local_xls}")

                        # Convert and upload
                        xlsx_path = convert_xls_to_xlsx(local_xls)
                        if xlsx_path and xlsx_path.exists():
                            # Ensure the original XLS file is deleted before uploading
                            if local_xls and local_xls.exists():
                                local_xls.unlink()
                                print(f"[Daily] 🗑 Deleted original XLS: {local_xls}")

                            # Now upload only the XLSX file
                            await upload_to_drive(xlsx_path, FOLDER_ID)
                            xlsx_path.unlink()
                            print(f"[Daily] 🗑 Deleted uploaded XLSX: {xlsx_path}")

                    else:
                        print(f"[Daily] No Excel icon found for {fac_name}")

                except Exception as ex_fac:
                    print(f"[Daily] Error processing {fac_name}: {ex_fac}")
                    traceback.print_exc()

        finally:
            await browser.close()
            print("[Daily] Browser closed")


def in_run_hours(now: datetime, start_hour: int, end_hour: int) -> bool:
    return (start_hour <= now.hour < end_hour)

def main():
    print(f"[Daily] Starting script with config: START_HOUR={START_HOUR}, END_HOUR={END_HOUR}, FREQ={FREQUENCY_MINUTES}min, FOLDER_ID={FOLDER_ID}")

    est_tz = pytz.timezone("US/Eastern")
    while True:
        try:
            now = datetime.now(est_tz)
            if in_run_hours(now, START_HOUR, END_HOUR):
                print(f"[Daily] Within run hours ({START_HOUR}-{END_HOUR}), running iteration...")
                start_time = time.time()
                asyncio.run(run_daily_iteration())
                elapsed = time.time() - start_time
                # Sleep until next frequency
                to_sleep = max(0, FREQUENCY_MINUTES*60 - elapsed)
                print(f"[Daily] Sleeping {to_sleep:.0f}s until next iteration.")
                time.sleep(to_sleep)
            else:
                print(f"[Daily] Outside run hours ({START_HOUR}-{END_HOUR}), waiting for next day...")
                next_run = now.replace(hour=START_HOUR, minute=0, second=0, microsecond=0)
                if next_run < now:
                    next_run += timedelta(days=1)
                sleep_secs = (next_run - now).total_seconds()
                print(f"[Daily] Sleeping {sleep_secs/3600:.1f}h until {next_run}...")
                time.sleep(sleep_secs)

        except KeyboardInterrupt:
            print("[Daily] KeyboardInterrupt => exiting.")
            sys.exit(0)
        except Exception as e:
            print(f"[Daily] Unexpected main-loop error: {e}")
            traceback.print_exc()
            time.sleep(30)  # short recovery delay

if __name__ == "__main__":
    main()
