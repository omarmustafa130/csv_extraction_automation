import asyncio
from playwright.async_api import async_playwright
from pathlib import Path
import time
from datetime import datetime, timedelta
import pytz
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import os
import pandas as pd
import traceback
from flask import Flask, render_template, request, jsonify
import threading
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from xlrd import open_workbook
from pathlib import Path
import traceback
import sys
app = Flask(__name__)

# Global control variables
script_running = False
control_lock = threading.Lock()
current_status = "Stopped"
run_hours = (9, 22)  # Start hour, End hour (24h format)
run_frequency = 60  # Minutes
script_username = os.getenv("SCRIPT_USERNAME", "Default_USERNAME")
script_password = os.getenv("SCRIPT_PASSWORD", "DEFAULT_PASSWORD")
folder_id = 'GOOGLE DRIVE FOLDER ID'

# Existing utility functions
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from xlrd import open_workbook
import os
# Fix Unicode error in Windows terminal
sys.stdout.reconfigure(encoding='utf-8')

def convert_xls_to_xlsx(xls_path) -> Path:
    """Convert an .xls file to .xlsx format while preserving merged formatting."""
    try:
        xls_path = Path(xls_path)  # Ensure xls_path is a Path object

        if not xls_path.exists():
            print(f"❌ File not found: {xls_path}")
            return None  # Avoid processing non-existent file

        xlsx_path = xls_path.with_suffix(".xlsx")

        print(f"🔄 Converting {xls_path} to {xlsx_path}...")

        # Read .xls file using xlrd
        xls_book = open_workbook(str(xls_path), formatting_info=True)
        xls_sheet = xls_book.sheet_by_index(0)

        # Create a new .xlsx workbook
        xlsx_book = Workbook()
        xlsx_sheet: Worksheet = xlsx_book.active

        # Copy data from .xls to .xlsx
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                xlsx_sheet.cell(row=row + 1, column=col + 1, value=xls_sheet.cell_value(row, col))

        # Preserve merged cells
        for merged_range in xls_sheet.merged_cells:
            r1, r2, c1, c2 = merged_range
            merge_range = f"{get_column_letter(c1 + 1)}{r1 + 1}:{get_column_letter(c2)}{r2}"
            xlsx_sheet.merge_cells(merge_range)

        # Save new .xlsx file
        xlsx_book.save(str(xlsx_path))
        print(f"✅ Successfully converted {xls_path} to {xlsx_path}")
        return xlsx_path

    except Exception as e:
        print(f"❌ Error converting {xls_path} to XLSX: {e}")
        traceback.print_exc()
        return None


def authenticate_drive():
    creds = service_account.Credentials.from_service_account_file(
        "SERVICE_ACCOUNT_JSON_FILE.json", 
        scopes=["https://www.googleapis.com/auth/drive", "https://www.googleapis.com/auth/spreadsheets"]
    )
    return build("drive", "v3", credentials=creds)

async def upload_to_drive(file_path: Path, folder_id: str):
    """Upload an XLSX file to Google Drive."""
    try:
        drive_service = authenticate_drive()
        file_metadata = {
            "name": file_path.name,
            "parents": [folder_id],
        }
        media = MediaFileUpload(
            str(file_path), 
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        uploaded_file = drive_service.files().create(
            body=file_metadata, media_body=media, fields="id"
        ).execute()
        print(f"✅ Uploaded {file_path} to Google Drive with ID: {uploaded_file['id']}")
    except Exception as e:
        print(f"❌ Error uploading {file_path}: {e}")
        traceback.print_exc()

async def rename_csv(file_path: Path, facility_name: str):
    """Renames the downloaded file to include timestamp and facility name."""
    est_tz = pytz.timezone('US/Eastern')
    current_time_est = datetime.now(est_tz)
    timestamp = current_time_est.strftime("%Y%m%d_%H%M%S")
    new_name = f"{timestamp}_{facility_name}.xls"
    new_path = file_path.parent / new_name
    try:
        file_path.rename(new_path)
        print(f"Renamed file to: {new_path}")
        return new_path
    except Exception as e:
        print(f"Error renaming file: {e}")
        return None

# Main automation function
async def automation_function():
    est_tz = pytz.timezone('US/Eastern')
    current_time_est = datetime.now(est_tz)
    new_date = current_time_est.strftime("%m/%d/%Y")
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()

        try:
            # Login sequence
            await page.goto("https://mybizaccount.fedex.com/")
            await page.click('input.credentials_input_submit')
            await page.wait_for_selector('#input28')
            await page.fill('#input28', script_username)
            await page.fill('#input36', script_password)
            await page.click('input.button-primary')
            await page.wait_for_load_state('networkidle')
            time.sleep(5)
            # Search navigation
            await page.fill('input#PTSKEYWORD', 'Daily Service Wk & Vision IBPR')
            await page.click('#PTSSEARCHBTN')
            await page.wait_for_load_state('networkidle')
            await page.keyboard.press("Enter")
            await page.wait_for_load_state('networkidle')

            # Date input
            await page.fill('div#dateTimePicker1 input', new_date)
            
            facilities = [
                {'name': 'ZECA-278', 'steps': ["Tab", "Enter", "ArrowUp", "Enter"]},
                {'name': 'ZWLN-256/3256', 'steps': ["Tab", "Enter", "ArrowDown", "ArrowDown", "Enter"]}
            ]

            for facility in facilities:
                print(f"Processing: {facility['name']}")
                await page.fill('div#dateTimePicker1 input', new_date)
                
                # Execute navigation steps
                for step in facility['steps']:
                    await page.keyboard.press(step)
                    await asyncio.sleep(0.5)

                try:
                    await page.click('button.selectionButton')
                    await page.wait_for_load_state('networkidle', timeout=60000)

                    excel_locator = page.locator("img[alt='Excel']")
                    if await excel_locator.count() > 0:
                        async with page.expect_download() as download_info:
                            await excel_locator.click()
                        download = await download_info.value

                        safe_name = facility['name'].replace('/', '_').replace(' ', '_')
                        original_path = Path.cwd() / f"{safe_name}.xls"
                        await download.save_as(original_path)
                        print(f"Downloaded {original_path}")

                        renamed_path = await rename_csv(original_path, safe_name)
                        xlsx_path = convert_xls_to_xlsx(renamed_path)
                        if xlsx_path:
                            await upload_to_drive(xlsx_path, folder_id)
                        else:
                            print(f"❌ Failed to convert file: {original_path}")
                    else:
                        print(f"No Excel icon found for {facility['name']}")
                except Exception as e:
                    print(f"Error processing {facility['name']}: {str(e)}")
                    traceback.print_exc()

        finally:
            await browser.close()

async def automation_loop():
    global script_running, current_status
    est_tz = pytz.timezone("US/Eastern")
    
    while script_running:
        now = datetime.now(est_tz)
        current_hour = now.hour
        
        if run_hours[0] <= current_hour < run_hours[1]:
            current_status = "Running"
            print(f"🟢 Running script at {now.strftime('%Y-%m-%d %I:%M %p EST')}")
            await automation_function()
            
            sleep_time = run_frequency * 60
            print(f"⏳ Next run in {run_frequency} minutes...")
            for _ in range(sleep_time):
                if not script_running:
                    break
                await asyncio.sleep(1)
        else:
            current_status = "Waiting"
            next_run = now.replace(hour=run_hours[0], minute=0, second=0, microsecond=0)
            if next_run < now:
                next_run += timedelta(days=1)
            sleep_time = (next_run - now).total_seconds()
            print(f"⏸️ Off-hours! Sleeping until {next_run.strftime('%I:%M %p EST')}")
            for _ in range(int(sleep_time)):
                if not script_running:
                    break
                await asyncio.sleep(1)

# Web Interface Endpoints
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/status')
def status():
    return jsonify({
        'running': script_running,
        'status': current_status,
        'hours': f"{run_hours[0]}:00 - {run_hours[1]}:00",
        'frequency': f"{run_frequency} minutes",
        'username': script_username
    })

@app.route('/control', methods=['POST'])
def control():
    global script_running
    action = request.json.get('action')
    
    with control_lock:
        if action == 'start' and not script_running:
            script_running = True
            threading.Thread(target=lambda: asyncio.run(automation_loop())).start()
            return jsonify({'status': 'starting'})
        elif action == 'stop' and script_running:
            script_running = False
            return jsonify({'status': 'stopping'})
        return jsonify({'status': 'no change'})

@app.route('/update_settings', methods=['POST'])
def update_settings():
    global run_hours, run_frequency, script_username, script_password
    data = request.json
    
    with control_lock:
        if 'start_hour' in data and 'end_hour' in data:
            run_hours = (int(data['start_hour']), int(data['end_hour']))
        if 'frequency' in data:
            run_frequency = int(data['frequency'])
        if 'username' in data:
            script_username = data['username']
        if 'password' in data:
            script_password = data['password']
            
    return jsonify({'status': 'settings updated'})

def run_flask():
    app.run(host='0.0.0.0', port=5000, use_reloader=False)

if __name__ == "__main__":
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.daemon = True
    flask_thread.start()
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        script_running = False