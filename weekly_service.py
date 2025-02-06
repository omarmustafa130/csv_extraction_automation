import os
import sys
import time
import traceback
import asyncio
import pytz
from datetime import datetime, timedelta
from pathlib import Path

from playwright.async_api import async_playwright
from xlrd import open_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

sys.stdout.reconfigure(encoding='utf-8')

# Load configuration from environment
FOLDER_ID = os.getenv("FOLDER_ID", "FOLDER_ID")
SCRIPT_USERNAME = os.getenv("SCRIPT_USERNAME", "USERNAME")
SCRIPT_PASSWORD = os.getenv("SCRIPT_PASSWORD", "PASSWORD")
SCHEDULE_RUN = int(os.getenv("SCHEDULE_RUN", "1"))  # 1 = Scheduled, 0 = Force Run Now

FACILITIES = [
    {"name": "ZNHI-250/3250", "steps": []},
    {"name": "ZECA-278",      "steps": ["Tab", "Enter", "ArrowUp", "Enter"]},
    {"name": "ZWLN-256/3256", "steps": ["Tab", "Enter", "ArrowDown", "ArrowDown", "Enter"]},
]

def authenticate_drive():
    creds = service_account.Credentials.from_service_account_file(
        "SERVICE_ACCOUNT_JSON.json",
        scopes=["https://www.googleapis.com/auth/drive", "https://www.googleapis.com/auth/spreadsheets"]
    )
    return build("drive", "v3", credentials=creds)

async def upload_to_drive(file_path: Path, folder_id: str):
    """Upload the file to Google Drive"""
    try:
        drive_service = authenticate_drive()
        file_metadata = {"name": file_path.name, "parents": [folder_id]}
        media = MediaFileUpload(str(file_path), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        uploaded_file = drive_service.files().create(body=file_metadata, media_body=media, fields="id").execute()
        print(f"[Weekly] ✅ Uploaded {file_path} (ID: {uploaded_file['id']})")
    except Exception as e:
        print(f"[Weekly] ❌ Error uploading {file_path}: {e}")
        traceback.print_exc()

def convert_xls_to_xlsx(xls_path: Path) -> Path:
    """Convert .xls to .xlsx while preserving merges."""
    try:
        if not xls_path.exists():
            print(f"[Weekly] ❌ File not found: {xls_path}")
            return None

        xlsx_path = xls_path.with_suffix(".xlsx")
        xls_book = open_workbook(str(xls_path), formatting_info=True)
        xlsx_book = Workbook()
        xlsx_book.remove(xlsx_book.active)

        for sheet_name in xls_book.sheet_names():
            xls_sheet = xls_book.sheet_by_name(sheet_name)
            new_sheet = xlsx_book.create_sheet(title=sheet_name)
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    new_sheet.cell(row=row_idx+1, column=col_idx+1, value=xls_sheet.cell_value(row_idx, col_idx))
            for (r1, r2, c1, c2) in xls_sheet.merged_cells:
                merge_range = f"{get_column_letter(c1+1)}{r1+1}:{get_column_letter(c2)}{r2}"
                new_sheet.merge_cells(merge_range)

        xlsx_book.save(str(xlsx_path))
        return xlsx_path
    except Exception as e:
        print(f"[Weekly] ❌ XLS->XLSX conversion failed: {e}")
        traceback.print_exc()
        return None

async def run_weekly_service():
    """Run the Playwright automation script."""
    est_tz = pytz.timezone("US/Eastern")
    new_date = datetime.now(est_tz).strftime("%m/%d/%Y")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()

        try:
            # Login sequence
            await page.goto("https://mybizaccount.fedex.com/")
            time.sleep(5)
            await page.click('input.credentials_input_submit')
            time.sleep(5)
            await page.wait_for_selector('#input28')
            await page.fill('#input28', SCRIPT_USERNAME)
            await page.fill('#input36', SCRIPT_PASSWORD)
            await page.click('input.button-primary')
            # Wait for 'domcontentloaded'
            await page.wait_for_load_state('domcontentloaded', timeout=60000)
            time.sleep(10)  # Consider replacing with an async sleep if possible
            # Search navigation
            await page.fill('input#PTSKEYWORD', 'Daily Service Wk & Vision IBPR')
            await page.click('#PTSSEARCHBTN')
            await page.wait_for_load_state('networkidle')
            time.sleep(5)
            await page.keyboard.press("Enter")

            await page.wait_for_load_state('networkidle')
            time.sleep(5)
            # Set the date
            await page.fill('div#dateTimePicker1 input', new_date)
            time.sleep(15)
            await page.click("li.triggered a[href*='/mgba/wsw']")
            await page.wait_for_load_state('networkidle')

            for fac in FACILITIES:
                fac_name = fac["name"]
                print(f"[Weekly] Processing facility: {fac_name}")
                await page.fill('div#dateTimePicker1 input', new_date)

                for key_step in fac["steps"]:
                    await page.keyboard.press(key_step)
                    await asyncio.sleep(0.5)

                try:
                    await page.click('button.selectionButton')
                    await page.wait_for_load_state('networkidle', timeout=60000)
                    excel_locator = page.locator("img[alt='Excel']")
                    if (await excel_locator.count()) > 0:
                        async with page.expect_download() as dl_info:
                            await excel_locator.click()
                        dl = await dl_info.value

                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        safe_name = fac_name.replace("/", "_").replace(" ", "_")
                        local_xls = Path.cwd() / f"{timestamp}_{safe_name}.xls"
                        await dl.save_as(local_xls)

                        xlsx_path = convert_xls_to_xlsx(local_xls)
                        if xlsx_path:
                            await upload_to_drive(xlsx_path, FOLDER_ID)
                            if local_xls.exists():
                                local_xls.unlink()
                            if xlsx_path.exists():
                                xlsx_path.unlink()
                except Exception as ex_fac:
                    print(f"[Weekly] Error processing {fac_name}: {ex_fac}")
                    traceback.print_exc()
        finally:
            await browser.close()
            print("[Weekly] Browser closed")

def should_run_today():
    """Check if today is Friday at 10 PM."""
    est_tz = pytz.timezone("US/Eastern")
    now = datetime.now(est_tz)
    return now.weekday() == 4 and now.hour == 22  # Friday at 10PM

def main():
    """Run the script based on schedule_run variable."""
    if SCHEDULE_RUN == 1:
        if should_run_today():
            print("[Weekly] Scheduled run: Executing...")
            asyncio.run(run_weekly_service())
        else:
            print("[Weekly] Not scheduled time. Run manually if needed.")
    else:
        print("[Weekly] Force run: Executing now...")
        asyncio.run(run_weekly_service())

if __name__ == "__main__":
    main()
